import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 1. Solicitar arquivo Excel via input
caminho_excel = input("🔍 Digite o caminho completo do arquivo Excel de RECEBIMENTOS: ").strip('"')
if not os.path.isfile(caminho_excel):
    print("❌ Arquivo não encontrado.")
    exit()

# 2. Pegar pasta do script (não da planilha)
pasta_script = os.path.dirname(os.path.abspath(__file__))  # Caminho onde está o script

# 3. Criar caminho para a pasta NFS-E relativa ao script
pasta_nfs = os.path.join(pasta_script, 'NFS-E')

# 4. Nome base do arquivo Excel
nome_base = os.path.splitext(os.path.basename(caminho_excel))[0]

# 3. Ler o Excel
xls = pd.ExcelFile(caminho_excel, engine='xlrd')
df = pd.concat([xls.parse(sheet_name, header=1) for sheet_name in xls.sheet_names], ignore_index=True)

# 4. Filtrar colunas úteis
colunas_certas = [
    'Data Pgto', 'Documento', 'Cliente', 'Valor Bruto', 'ISS', 'INSS',
    'PIS', 'COFINS', 'IR', 'CSLL', 'JUROS', 'DESCONTO', 'TX ADM',
    'Valor Líquido', 'Valor Recebido', 'OBS'
]
df = df[colunas_certas]

# 5. Filtrar apenas documentos com padrão de NF
padrao_nf = re.compile(r'NF \d{4}/\d+')
df = df[df['Documento'].astype(str).str.contains(padrao_nf)]

# 6. Transformar Documento em número de 15 dígitos
def formatar_nf(documento):
    match = re.search(r'NF (\d{4})/(\d+)', str(documento), re.IGNORECASE)
    if match:
        ano = match.group(1)
        numero = match.group(2).zfill(11)
        return f"{ano}{numero}"
    return documento

df['Documento'] = df['Documento'].apply(formatar_nf)
df['Ano'] = df['Documento'].str[:4]
df['CNPJ'] = ''
df['Data Vencimento'] = ''

# 7. Consultar os CSVs de NFS-E por ano
for idx, row in df.iterrows():
    ano = row['Ano']
    nota = row['Documento']
    caminho_csv = os.path.join(pasta_nfs, f'NFS-E {ano}.csv')
    if not os.path.exists(caminho_csv):
        continue

    try:
        df_csv = pd.read_csv(caminho_csv, encoding='utf-8-sig', sep=';')
        col_numero = next((c for c in df_csv.columns if 'ÚMERO' in c.upper()), None)
        col_tomador = next((c for c in df_csv.columns if 'TOMADOR' in c.upper()), None)
        col_emissao = next((c for c in df_csv.columns if re.search(r'emiss[aã][oõ]', c, re.IGNORECASE)), None)
        if not col_numero or not col_tomador or not col_emissao:
            continue

        df_csv['NumeroFormatado'] = df_csv[col_numero].astype(str).str.replace(r'\.0$', '', regex=True).str.zfill(15)
        resultado = df_csv[df_csv['NumeroFormatado'] == nota]

        if not resultado.empty:
            df.at[idx, 'CNPJ'] = resultado.iloc[0][col_tomador]
            emissao_raw = resultado.iloc[0][col_emissao]
            emissao_data = pd.to_datetime(str(emissao_raw), dayfirst=True, errors='coerce')
            if pd.notna(emissao_data):
                vencimento = emissao_data + pd.Timedelta(days=30)
                df.at[idx, 'Data Vencimento'] = vencimento.strftime('%d/%m/%Y')
    except:
        continue

# 8. Organizar e preparar colunas
df['Valor Multa'] = 0.00
df = df.drop(columns=[col for col in df.columns if col.strip().upper() == 'CLIENTE'], errors='ignore')
df['Data Pgto'] = pd.to_datetime(df['Data Pgto'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')

colunas_desejadas = [
    'Documento', 'CNPJ', 'Data Vencimento', 'Data Pgto',
    'Valor Líquido', 'JUROS', 'Valor Multa',
    'DESCONTO', 'PIS', 'COFINS', 'CSLL', 'IR'
]
df = df[[col for col in colunas_desejadas if col in df.columns]]

df['CNPJ'] = df['CNPJ'].astype(str).str.replace(r'\D', '', regex=True)
colunas_numericas = ['Valor Líquido', 'JUROS', 'Valor Multa', 'DESCONTO', 'PIS', 'COFINS', 'CSLL', 'IR']
for col in colunas_numericas:
    if col in df.columns:
        df[col] = df[col].fillna(0)

# 9. Salvar Excel
saida_excel = os.path.join(pasta_script, f'IMPORTACAO__{nome_base}.xlsx')
if os.path.exists(saida_excel):
    os.remove(saida_excel)
df.to_excel(saida_excel, index=False)

# 10. Formatar Excel
wb = load_workbook(saida_excel)
ws = wb.active
for cell in ws[1]:
    cell.font = Font(bold=True)
for col in ws.iter_cols(min_row=2):
    titulo = ws[f"{col[0].column_letter}1"].value
    if titulo and "Data" in titulo:
        for cell in col:
            cell.number_format = 'DD/MM/YYYY'
    elif titulo == "Documento":
        for cell in col:
            cell.number_format = '0'
for col in ws.columns:
    largura = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    letra = get_column_letter(col[0].column)
    ws.column_dimensions[letra].width = largura + 2
ws.auto_filter.ref = ws.dimensions
wb.save(saida_excel)

# 11. Gerar TXT
saida_txt = os.path.join(pasta_script, f'IMPORTACAO__{nome_base}.txt')
df_txt = df.copy()
for col in df_txt.columns:
    if col not in ['Documento', 'CNPJ', 'Data Vencimento', 'Data Pgto']:
        df_txt[col] = df_txt[col].apply(
            lambda x: f"{int(x)}" if float(x) == int(x) else f"{x:.2f}".replace('.', ',')
        )
df_txt.to_csv(saida_txt, sep=';', index=False, header=False, encoding='utf-8-sig')

print(f"\n✅ Arquivo Excel salvo em:\n📁 {saida_excel}")
print(f"\n📝 Arquivo TXT gerado com sucesso em:\n📁 {saida_txt}")
