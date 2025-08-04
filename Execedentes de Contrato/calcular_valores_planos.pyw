import pandas as pd
import numpy as np
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers

# Faixas fixas declaradas no código
faixas_planos = {
    "Basico": [
        (20000, 4, 349.71),
        (35000, 6, 489.63),
        (100000, 10, 629.51),
        (200000, 20, 847.75)
    ],
    "Plus": [
        (20000, 4, 489.63),
        (35000, 6, 629.51),
        (100000, 10, 847.75),
        (200000, 15, 1133.13),
        (300000, 20, 1695.48)
    ],
    "Premium": [
        (20000, 4, 847.75),
        (35000, 6, 986.26),
        (100000, 10, 1681.49),
        (200000, 15, 2543.25),
        (300000, 20, 3390.96)
    ]
}
ordem_planos = ["Basico", "Plus", "Premium"]
valor_funcionario = 57.78

# Cálculo
def calcular(row):
    plano = row["Plano"]
    faturamento = row["Valor do Contratos - Faturado"]
    funcionarios = row["N° de Funcionarios - Faturado"]
    
    faixas = faixas_planos[plano]
    plano_recomendado = plano
    valor_base = faixas[-1][2]
    limite_func = faixas[-1][1]

    for limite_fat, limite_func_opt, valor in faixas:
        if faturamento <= limite_fat:
            valor_base = valor
            limite_func = limite_func_opt
            break
    else:
        idx = ordem_planos.index(plano)
        if idx + 1 < len(ordem_planos):
            plano_recomendado = ordem_planos[idx + 1]

    excedente = max(0, funcionarios - limite_func)
    adicional = excedente * valor_funcionario
    valor_total = valor_base + adicional

    return pd.Series([plano_recomendado, valor_base, limite_func, excedente, adicional, valor_total])

# Função principal
def processar():
    try:
        caminho_planos = filedialog.askopenfilename(title="Selecione o arquivo de planos")
        caminho_dados = filedialog.askopenfilename(title="Selecione o arquivo de dados de faturamento")
        
        if not caminho_planos or not caminho_dados:
            messagebox.showerror("Erro", "Você deve selecionar os dois arquivos.")
            return

        df_planos = pd.read_excel(caminho_planos)
        df_dados = pd.read_excel(caminho_dados)

        df_planos.columns = df_planos.columns.str.strip()
        df_dados.columns = df_dados.columns.str.strip()
        df_planos["Plano"] = df_planos["Plano"].str.strip().str.capitalize()

        df = pd.merge(df_dados, df_planos, on="Cód.", how="left")

        df[[ 
            "Plano Recomendado", "Valor Base", "Funcionários Permitidos", 
            "Excedente Funcionários", "Adicional Funcionários", "Valor Total" 
        ]] = df.apply(calcular, axis=1)

        df_saida = df[[ 
            "Cód.",
            "Cliente_x",
            "Plano",
            "Plano Recomendado",
            "Valor do Contrato - Contratado",
            "Valor do Contratos - Faturado",
            "N° de Funcionarios - Faturado",
            "Excedente Funcionários",
            "Valor Base",
            "Valor Total"
        ]].rename(columns={
            "Cliente_x": "Cliente",
            "Valor do Contrato - Contratado": "Valor do Contrato",
            "Valor do Contratos - Faturado": "Faturamento",
            "N° de Funcionarios - Faturado": "Funcionários",
            "Excedente Funcionários": "Excedente (funcionários)",
            "Valor Total": "Total"
        })

        saida = os.path.join(os.getcwd(), "Resultado_Final.xlsx")
        df_saida.to_excel(saida, index=False)

        # 🧩 Personalize as colunas:
        colunas_centralizar = [
            "Cód.",
            "Plano",
            "Plano Recomendado",
            "Funcionários",
            "Excedente (funcionários)"
        ]

        colunas_moeda = [
            "Valor do Contrato",
            "Faturamento",
            "Valor Base",
            "Total"
        ]

        # 📊 Abrir planilha e identificar colunas
        wb = load_workbook(saida)
        ws = wb.active
        cabecalhos = [cell.value for cell in ws[1]]

        # Mapear os índices das colunas selecionadas
        idx_centralizar = [i+1 for i, col in enumerate(cabecalhos) if col in colunas_centralizar]
        idx_moeda = [i+1 for i, col in enumerate(cabecalhos) if col in colunas_moeda]

        # Aplicar estilos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for i, cell in enumerate(row, start=1):
                if i in idx_centralizar:
                    cell.alignment = Alignment(horizontal="center")
                if i in idx_moeda:
                    cell.number_format = 'R$ #,##0.00'

        # 💾 Salvar alterações
        wb.save(saida)

        messagebox.showinfo("Sucesso", f"✅ Resultado salvo em:\n{saida}")
    except Exception as e:
        messagebox.showerror("Erro durante o processamento", str(e))

# GUI
janela = tk.Tk()
janela.title("Calculadora de Planos")
janela.geometry("400x200")

label = tk.Label(janela, text="Clique no botão abaixo para iniciar", font=("Arial", 12))
label.pack(pady=20)

botao = tk.Button(janela, text="Selecionar Arquivos e Processar", command=processar, font=("Arial", 11), bg="#4CAF50", fg="white", padx=10, pady=5)
botao.pack()

janela.mainloop()
